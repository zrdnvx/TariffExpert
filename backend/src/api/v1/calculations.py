from uuid import UUID
from io import BytesIO
from datetime import timezone
from decimal import Decimal
import re
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from src.api.services.calculator import TariffCalculatorService
from src.db.models.models import Building, Calculation, CalculationItem
from src.schemas.calculation import (
    CalculationResponse,
    CalculationHistoryItemResponse,
    CalculationDetailResponse,
    CalculationUpdateRequest,
    DetailCalculationRequest,
    DetailCalculationResponse,
    DetailComponentResponse,
)
from src.db.database import get_async_session
from src.api.v1.auth import get_current_user, CurrentUser

router = APIRouter(
    prefix="/calculations",
    tags=["Calculations"],
    dependencies=[Depends(get_current_user)],
)


def _safe_filename_part(value: str, max_len: int = 80) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|]+", "", value).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned[:max_len] if cleaned else "dom"


def _safe_ascii_filename(value: str, max_len: int = 100) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-")
    if not cleaned:
        return "calculation"
    return cleaned[:max_len]


def _outer_border_for_cell(row: int, col: str, top: int, bottom: int, left_col: str, right_col: str) -> Border:
    return Border(
        left=Side(style="medium" if col == left_col else "thin", color="000000"),
        right=Side(style="medium" if col == right_col else "thin", color="000000"),
        top=Side(style="medium" if row == top else "thin", color="000000"),
        bottom=Side(style="medium" if row == bottom else "thin", color="000000"),
    )


async def _build_items_with_formulas(
    calc: Calculation, db: AsyncSession
):
    service = TariffCalculatorService(db)
    building = calc.building
    k1 = await service._resolve_k1(building) if building else None
    k2 = await service._resolve_k2(building) if building else None
    result = []
    for item in calc.items:
        formula_label = None
        formula_substitution = None
        base_rate = Decimal(item.tariff.rate) if item.tariff else Decimal(item.applied_rate)
        applied = Decimal(item.applied_rate)
        item_number = (item.item_number or "").rstrip(".")
        if (
            item_number == TariffCalculatorService.REPAIR_BASE_ITEM_NUMBER.rstrip(".")
            and k1
            and k2
        ):
            formula_label = "Птрк = БС × К1 × К2"
            formula_substitution = f"{base_rate:.4f} × {k1:.4f} × {k2:.4f} = {applied:.4f}"
        elif item_number == TariffCalculatorService.LANDSCAPING_BASE_ITEM_NUMBER.rstrip(".") and k2:
            formula_label = "Пбпт = БС × К2"
            formula_substitution = f"{base_rate:.4f} × {k2:.4f} = {applied:.4f}"
        result.append(
            {
                "id": item.id,
                "item_number": item.item_number,
                "name": item.name,
                "applied_rate": applied,
                "formula_label": formula_label,
                "formula_substitution": formula_substitution,
            }
        )
    return result


@router.post("/{building_id}/run", response_model=CalculationResponse)
async def run_calculation(
    building_id: UUID,
    db: AsyncSession = Depends(get_async_session),
    user: CurrentUser = Depends(get_current_user),
):
    # 1. Находим здание
    building = await db.get(Building, building_id)
    if not building:
        raise HTTPException(status_code=404, detail="Building not found")
    if building.organization_id != user.organization_id:
        raise HTTPException(status_code=404, detail="Building not found")

    # 2. Считаем (режим «Расчет размера платы…»)
    service = TariffCalculatorService(db)
    result = await service.calculate(building)

    # 3. Подгружаем связи для корректного ответа (детализация строк)
    stmt = (
        select(Calculation)
        .options(selectinload(Calculation.items).joinedload(CalculationItem.tariff))
        .where(Calculation.id == result.id)
    )
    result_with_items = await db.execute(stmt)
    calc = result_with_items.scalar_one()
    items = await _build_items_with_formulas(calc, db)
    return CalculationResponse(
        id=calc.id,
        building_id=calc.building_id,
        total_rate=calc.total_rate,
        items=items,
    )


@router.get("/", response_model=list[CalculationHistoryItemResponse])
async def list_calculations(
    db: AsyncSession = Depends(get_async_session),
    user: CurrentUser = Depends(get_current_user),
):
    stmt = (
        select(Calculation, Building.address)
        .join(Building, Building.id == Calculation.building_id)
        .where(Calculation.organization_id == user.organization_id)
        .order_by(Calculation.created_at.desc())
    )
    result = await db.execute(stmt)
    rows = result.all()
    return [
        CalculationHistoryItemResponse(
            id=calc.id,
            created_at=calc.created_at,
            building_id=calc.building_id,
            building_address=address,
            total_rate=calc.total_rate,
        )
        for calc, address in rows
    ]


@router.get("/{calculation_id}", response_model=CalculationDetailResponse)
async def get_calculation(
    calculation_id: UUID,
    db: AsyncSession = Depends(get_async_session),
    user: CurrentUser = Depends(get_current_user),
):
    stmt = (
        select(Calculation)
        .options(
            selectinload(Calculation.items).joinedload(CalculationItem.tariff),
            selectinload(Calculation.building),
        )
        .where(Calculation.id == calculation_id)
    )
    result = await db.execute(stmt)
    calc = result.scalar_one_or_none()
    if not calc:
        raise HTTPException(status_code=404, detail="Calculation not found")
    if calc.organization_id != user.organization_id and user.role != "superadmin":
        raise HTTPException(status_code=404, detail="Calculation not found")

    items = await _build_items_with_formulas(calc, db)
    return CalculationDetailResponse(
        id=calc.id,
        building_id=calc.building_id,
        building_address=calc.building.address if calc.building else "—",
        created_at=calc.created_at,
        total_rate=calc.total_rate,
        items=items,
    )


@router.patch("/{calculation_id}", response_model=CalculationDetailResponse)
async def update_calculation(
    calculation_id: UUID,
    payload: CalculationUpdateRequest,
    db: AsyncSession = Depends(get_async_session),
    user: CurrentUser = Depends(get_current_user),
):
    stmt = (
        select(Calculation)
        .options(
            selectinload(Calculation.items).joinedload(CalculationItem.tariff),
            selectinload(Calculation.building),
        )
        .where(Calculation.id == calculation_id)
    )
    result = await db.execute(stmt)
    calc = result.scalar_one_or_none()
    if not calc:
        raise HTTPException(status_code=404, detail="Calculation not found")
    if calc.organization_id != user.organization_id and user.role != "superadmin":
        raise HTTPException(status_code=404, detail="Calculation not found")

    items_by_id = {item.id: item for item in calc.items}
    for upd in payload.items:
        item = items_by_id.get(upd.item_id)
        if not item:
            raise HTTPException(status_code=400, detail=f"Calculation item not found: {upd.item_id}")
        item.applied_rate = upd.applied_rate

    calc.total_rate = sum((Decimal(item.applied_rate) for item in calc.items), Decimal("0"))
    await db.commit()
    await db.refresh(calc)

    refreshed_stmt = (
        select(Calculation)
        .options(
            selectinload(Calculation.items).joinedload(CalculationItem.tariff),
            selectinload(Calculation.building),
        )
        .where(Calculation.id == calc.id)
    )
    refreshed_res = await db.execute(refreshed_stmt)
    updated = refreshed_res.scalar_one()

    items = await _build_items_with_formulas(updated, db)
    return CalculationDetailResponse(
        id=updated.id,
        building_id=updated.building_id,
        building_address=updated.building.address if updated.building else "—",
        created_at=updated.created_at,
        total_rate=updated.total_rate,
        items=items,
    )


@router.delete("/{calculation_id}")
async def delete_calculation(
    calculation_id: UUID,
    db: AsyncSession = Depends(get_async_session),
    user: CurrentUser = Depends(get_current_user),
):
    calc = await db.get(Calculation, calculation_id)
    if not calc:
        raise HTTPException(status_code=404, detail="Calculation not found")
    if calc.organization_id != user.organization_id and user.role != "superadmin":
        raise HTTPException(status_code=404, detail="Calculation not found")

    await db.execute(
        delete(CalculationItem).where(CalculationItem.calculation_id == calculation_id)
    )
    await db.execute(delete(Calculation).where(Calculation.id == calculation_id))
    await db.commit()
    return {"status": "ok"}


@router.post("/detail", response_model=DetailCalculationResponse)
async def detail_calculation(
    payload: DetailCalculationRequest,
    db: AsyncSession = Depends(get_async_session),
    user: CurrentUser = Depends(get_current_user),
):
    """
    Режим «Детализация платы…» (ТЗ, п. 2.3).
    На вход: адрес, существующая плата, площадь, этажность, год ввода и чекбоксы систем.
    На выход: разложение заданной платы по видам работ приложения 1.
    """
    service = TariffCalculatorService(db)
    # для подбора тарифов нужен city_id; берём из токена
    building_like = payload.model_copy(update={"city_id": user.city_id})
    breakdown = await service.build_detail_breakdown(
        building_like=building_like,
        existing_rate=payload.existing_rate,
    )

    components = [
        DetailComponentResponse(**c) for c in breakdown["components"]
    ]

    return DetailCalculationResponse(
        address=payload.address,
        existing_rate=payload.existing_rate,
        total_normative_rate=breakdown["total_normative_rate"],
        components=components,
    )


@router.get("/{calculation_id}/export")
async def export_calculation_to_excel(
    calculation_id: UUID,
    db: AsyncSession = Depends(get_async_session),
    user: CurrentUser = Depends(get_current_user),
):
    """
    Выгрузка результатов расчета (режим «Расчет размера платы…») в формате Excel.
    """
    stmt = (
        select(Calculation)
        .options(
            selectinload(Calculation.items).joinedload(CalculationItem.tariff),
            selectinload(Calculation.building),
        )
        .where(Calculation.id == calculation_id)
    )
    result = await db.execute(stmt)
    calc = result.scalar_one_or_none()

    if not calc:
        raise HTTPException(status_code=404, detail="Calculation not found")
    if calc.organization_id != user.organization_id and user.role != "superadmin":
        raise HTTPException(status_code=404, detail="Calculation not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Перечень работ"

    # Монохромный стиль под официальный бланк
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    header_fill = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
    title_font = Font(color="000000", bold=True, size=12)
    header_font = Font(color="000000", bold=True, size=10)
    normal_font = Font(color="000000", size=10)
    total_font = Font(color="000000", bold=True, size=11)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 78
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 44

    building = calc.building
    total_area = float(building.total_area) if building else 0.0
    created_dt_utc = calc.created_at.astimezone(timezone.utc)
    created_date_str = created_dt_utc.strftime("%d.%m.%Y")
    created_time_str = created_dt_utc.strftime("%H:%M UTC")
    created_date_for_filename = calc.created_at.astimezone(timezone.utc).strftime("%Y-%m-%d")
    total_amount = float(calc.total_rate) * total_area
    service = TariffCalculatorService(db)
    k1 = await service._resolve_k1(building) if building else None
    k2 = await service._resolve_k2(building) if building else None
    house_type_labels = {
        "monolith_brick": "Монолитные/кирпичные стены",
        "reinforced_concrete": "Железобетонные стены",
        "other_low_capital": "Пониженная капитальность, прочие материалы",
    }
    house_type_value = getattr(building, "house_type", None)
    house_type_key = house_type_value.value if hasattr(house_type_value, "value") else house_type_value

    # Заголовок как в образце
    ws.merge_cells("A1:E1")
    ws["A1"] = "П Е Р Е Ч Е Н Ь"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A2:E2")
    ws["A2"] = "работ и услуг по содержанию и ремонту общего имущества"
    ws["A2"].font = Font(bold=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A3:E3")
    ws["A3"] = (
        f"в многоквартирном доме по адресу: {building.address if building else '—'}"
    )
    ws["A3"].alignment = Alignment(horizontal="center")
    ws["A3"].font = Font(size=10)

    # Блок информации о доме и расчете (правый, с рамкой)
    ws.merge_cells("C5:D5")
    ws["C5"] = "Сведения о доме и расчете"
    ws["C5"].font = header_font
    ws["C5"].alignment = Alignment(horizontal="center")
    ws["C5"].fill = header_fill

    yes_no = lambda v: "Да" if v else "Нет"
    info_rows = [
        ("Дата расчета", created_date_str),
        ("Время расчета", created_time_str),
        ("Адрес дома", building.address if building else "—"),
        ("Общая площадь дома, кв.м", total_area),
        ("Этажность дома", building.floors_count if building else "—"),
        ("Год ввода в эксплуатацию", building.year_built if building else "—"),
        ("Тип объекта", "Многоквартирный дом" if getattr(building, "is_apartment_building", True) else "Прочий объект"),
        ("Тип многоквартирного дома (К1)", house_type_labels.get(house_type_key, "—") if getattr(building, "is_apartment_building", True) else "Не применяется"),
        ("Коэффициент К1", float(k1) if k1 is not None and getattr(building, "is_apartment_building", True) else "Не применяется"),
        ("Коэффициент К2", float(k2) if k2 is not None and getattr(building, "is_apartment_building", True) else "Не применяется"),
        ("Лифт", yes_no(getattr(building, "has_elevator", False))),
        ("Газ", yes_no(getattr(building, "has_gas", False))),
        ("Мусоропровод", yes_no(getattr(building, "has_trash_chute", False))),
        ("Центральное отопление", yes_no(getattr(building, "has_central_heating", False))),
        ("Локальная котельная", yes_no(getattr(building, "has_local_boiler", False))),
        ("Размер обязательного платежа, руб", total_amount),
        ("Размер обязательного платежа, руб/м²", float(calc.total_rate)),
    ]

    row = 6
    info_start_row = row
    for label, value in info_rows:
        ws[f"C{row}"] = label
        ws[f"C{row}"].font = header_font
        ws[f"C{row}"].alignment = Alignment(vertical="center", wrap_text=True)
        ws[f"D{row}"] = value
        ws[f"D{row}"].font = normal_font
        ws[f"D{row}"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        if isinstance(value, float):
            ws[f"D{row}"].number_format = "#,##0.00"
        ws.row_dimensions[row].height = 22
        row += 1
    info_end_row = row - 1

    # Рамка вокруг правого блока C5:D{info_end_row}
    for r in range(5, info_end_row + 1):
        for c in ("C", "D"):
            ws[f"{c}{r}"].border = _outer_border_for_cell(
                r, c, 5, info_end_row, "C", "D"
            )

    row += 1
    start_table_row = row
    ws[f"A{row}"] = "Наименование работ"
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"C{row}"] = "Плановая стоимость работ в год, руб"
    ws[f"D{row}"] = "Стоимость на 1 кв.м в месяц, руб"
    ws[f"E{row}"] = "Формула и подстановка"
    for col in ("A", "B", "C", "D", "E"):
        ws[f"{col}{row}"].font = header_font
        ws[f"{col}{row}"].fill = header_fill
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[f"{col}{row}"].border = thin_border
    ws.row_dimensions[row].height = 34

    row += 1
    sorted_items = sorted(calc.items, key=lambda i: i.item_number)
    for item in sorted_items:
        annual_amount = float(item.applied_rate) * total_area * 12.0
        ws[f"A{row}"] = item.item_number
        ws[f"B{row}"] = item.name
        ws[f"C{row}"] = annual_amount
        ws[f"D{row}"] = float(item.applied_rate)
        formula_text = "—"
        item_num = (item.item_number or "").rstrip(".")
        base_rate = Decimal(item.tariff.rate) if item.tariff else Decimal(item.applied_rate)
        applied_rate = Decimal(item.applied_rate)
        if item_num == TariffCalculatorService.REPAIR_BASE_ITEM_NUMBER.rstrip(".") and k1 is not None and k2 is not None:
            formula_text = (
                f"Птрк = БС × К1 × К2\n"
                f"{base_rate:.4f} × {Decimal(k1):.4f} × {Decimal(k2):.4f} = {applied_rate:.4f}"
            )
        elif item_num == TariffCalculatorService.LANDSCAPING_BASE_ITEM_NUMBER.rstrip(".") and k2 is not None:
            formula_text = (
                f"Пбпт = БС × К2\n"
                f"{base_rate:.4f} × {Decimal(k2):.4f} = {applied_rate:.4f}"
            )
        ws[f"E{row}"] = formula_text

        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="top")
        ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"C{row}"].alignment = Alignment(horizontal="right", vertical="top")
        ws[f"D{row}"].alignment = Alignment(horizontal="right", vertical="top")
        ws[f"E{row}"].alignment = Alignment(wrap_text=True, vertical="top")

        ws[f"A{row}"].font = normal_font
        ws[f"B{row}"].font = normal_font
        ws[f"C{row}"].font = normal_font
        ws[f"D{row}"].font = normal_font
        ws[f"E{row}"].font = normal_font
        ws[f"C{row}"].number_format = "#,##0.00"
        ws[f"D{row}"].number_format = "#,##0.0000"

        ws[f"A{row}"].border = thin_border
        ws[f"B{row}"].border = thin_border
        ws[f"C{row}"].border = thin_border
        ws[f"D{row}"].border = thin_border
        ws[f"E{row}"].border = thin_border
        row += 1

    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "Всего работ и услуг"
    ws[f"C{row}"] = total_amount
    ws[f"D{row}"] = float(calc.total_rate)
    ws[f"C{row}"].number_format = "#,##0.00"
    ws[f"D{row}"].number_format = "#,##0.0000"
    ws[f"A{row}"].font = total_font
    ws[f"C{row}"].font = total_font
    ws[f"D{row}"].font = total_font
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"C{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"D{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"A{row}"].border = thin_border
    ws[f"B{row}"].border = thin_border
    ws[f"C{row}"].border = thin_border
    ws[f"D{row}"].border = thin_border
    ws[f"E{row}"].border = thin_border

    # Высота строк
    for current_row in range(start_table_row + 1, row):
        cell_val = ws[f"B{current_row}"].value or ""
        lines = max(1, len(str(cell_val)) // 58 + 1)
        ws.row_dimensions[current_row].height = 16 * lines

    # Лист "Периодичность" в стиле образца
    ws_period = wb.create_sheet("Периодичность")
    ws_period.column_dimensions["A"].width = 98
    ws_period.column_dimensions["B"].width = 28

    ws_period.merge_cells("A1:B1")
    ws_period["A1"] = "Периодичность выполнения отдельных видов работ"
    ws_period["A1"].font = header_font
    ws_period["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws_period["A3"] = "Наименование работ"
    ws_period["B3"] = "Периодичность, раз в год"
    for col in ("A", "B"):
        ws_period[f"{col}3"].font = header_font
        ws_period[f"{col}3"].fill = header_fill
        ws_period[f"{col}3"].alignment = Alignment(horizontal="center", vertical="center")
        ws_period[f"{col}3"].border = thin_border
    ws_period.row_dimensions[3].height = 28

    # Базовые ориентировочные значения (можно уточнять регламентом УК)
    periodicity_defaults = {
        "2.4.1.": 52,
        "2.4.2.": 12,
        "2.4.3.": 180,
        "2.4.4.": 180,
        "2.4.5.": 2,
        "2.4.6.": 2,
        "2.4.7.": 26,
        "2.4.8.": 2,
        "2.4.9.": 26,
        "2.7.4": 1,
    }
    row_p = 4
    for item in sorted_items:
        ws_period[f"A{row_p}"] = f"{item.item_number} {item.name}"
        ws_period[f"B{row_p}"] = periodicity_defaults.get(item.item_number, "по регламенту")
        ws_period[f"A{row_p}"].font = normal_font
        ws_period[f"B{row_p}"].font = normal_font
        ws_period[f"A{row_p}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_period[f"B{row_p}"].alignment = Alignment(horizontal="center", vertical="top")
        ws_period[f"A{row_p}"].border = thin_border
        ws_period[f"B{row_p}"].border = thin_border
        text = ws_period[f"A{row_p}"].value or ""
        ws_period.row_dimensions[row_p].height = max(18, 16 * (len(str(text)) // 78 + 1))
        row_p += 1

    ws_period[f"A{row_p+1}"] = "Должность ____________________"
    ws_period[f"B{row_p+1}"] = "ФИО ____________________"
    ws_period[f"A{row_p+1}"].font = normal_font
    ws_period[f"B{row_p+1}"].font = normal_font

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    address_part = _safe_filename_part(building.address if building else "dom")
    filename_utf8 = f"raschet_{address_part}_{created_date_for_filename}.xlsx"
    filename_ascii = _safe_ascii_filename(filename_utf8)
    content_disposition = (
        f"attachment; filename=\"{filename_ascii}\"; "
        f"filename*=UTF-8''{quote(filename_utf8)}"
    )
    return StreamingResponse(
        stream,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": content_disposition,
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
